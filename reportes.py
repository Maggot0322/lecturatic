# -*- coding: utf-8 -*-
"""
reportes.py
------------------------------------------------------------
Genera reportes descargables para el panel del administrador
(PDF, Excel, CSV) y certificados en PDF para los estudiantes
que completan la plataforma.
------------------------------------------------------------
"""

import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.colors import HexColor


ENCABEZADOS = ["Nombre", "Usuario", "Promedio General", "Módulos Aprobados",
               "Porcentaje de Avance", "Último Acceso"]


def generar_csv(filas):
    """Genera un archivo CSV en memoria a partir de una lista de estudiantes."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(ENCABEZADOS)
    for f in filas:
        writer.writerow(f)
    return output.getvalue().encode("utf-8-sig")


def generar_excel(filas):
    """Genera un archivo Excel (.xlsx) en memoria a partir de una lista de estudiantes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte LECTURATIC"

    ws.append(ENCABEZADOS)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for f in filas:
        ws.append(f)

    for col in ws.columns:
        longitud = max(len(str(c.value)) for c in col if c.value is not None) if col else 10
        ws.column_dimensions[col[0].column_letter].width = max(15, longitud + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_pdf(filas, titulo="Reporte General de Estudiantes - LECTURATIC"):
    """Genera un reporte en PDF con tabla de estudiantes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle("titulo", parent=styles["Title"], alignment=TA_CENTER, fontSize=16)
    subtitulo_style = ParagraphStyle("subtitulo", parent=styles["Normal"], alignment=TA_CENTER, fontSize=10)

    elementos = [
        Paragraph(titulo, titulo_style),
        Paragraph(f"Generado el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitulo_style),
        Spacer(1, 0.6 * cm),
    ]

    data = [ENCABEZADOS] + [[str(x) for x in f] for f in filas]
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F8E9")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    buffer.seek(0)
    return buffer.getvalue()

def generar_certificado_pdf(nombre_estudiante, promedio_general):
    """
    Genera un certificado de logro más atractivo para estudiantes
    de quinto grado de primaria.
    """

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=1.2*cm,
        leftMargin=1.2*cm,
        topMargin=1.2*cm,
        bottomMargin=1.2*cm
    )

    styles = getSampleStyleSheet()

    titulo = ParagraphStyle(
        "Titulo",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=30,
        leading=34,
        textColor=HexColor("#2E7D32"),
        spaceAfter=10
    )

    subtitulo = ParagraphStyle(
        "Subtitulo",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=17,
        textColor=HexColor("#1565C0")
    )

    nombre = ParagraphStyle(
        "Nombre",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=28,
        leading=32,
        textColor=HexColor("#0D47A1"),
        spaceAfter=15
    )

    texto = ParagraphStyle(
        "Texto",
        parent=styles["BodyText"],
        alignment=TA_CENTER,
        fontSize=14,
        leading=22
    )

    firma = ParagraphStyle(
        "Firma",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=12
    )

    # Medalla según promedio

    if promedio_general >= 4.8:
        medalla = "✺ MEDALLA DE ORO"
    elif promedio_general >= 4:
        medalla = "❖ MEDALLA DE PLATA"
    else:
        medalla = "✵ MEDALLA DE BRONCE"

    elementos = []

    # Marco superior

    tabla = Table([[""]], colWidths=25*cm, rowHeights=0.45*cm)

    tabla.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),HexColor("#2E7D32")),
        ("BOX",(0,0),(-1,-1),3,HexColor("#1565C0"))
    ]))

    elementos.append(tabla)

    elementos.append(Spacer(1,0.5*cm))

    elementos.append(Paragraph("🏆 CERTIFICADO DE LOGRO 🏆",titulo))

    elementos.append(
        Paragraph(
            "LECTURATIC - Plataforma para el Fortalecimiento de la Comprensión Lectora",
            subtitulo
        )
    )

    elementos.append(Spacer(1,0.8*cm))

    elementos.append(
        Paragraph(
            "<b>Este certificado se otorga a:</b>",
            texto
        )
    )

    elementos.append(Spacer(1,0.3*cm))

    elementos.append(
        Paragraph(
            f"<b>{nombre_estudiante.upper()}</b>",
            nombre
        )
    )

    elementos.append(Spacer(0,0*cm))

    mensaje = f"""
    Por haber culminado satisfactoriamente los cinco módulos (5) del Entorno
    Virtual de Aprendizaje <b>LECTURATIC</b>, demostrando habilidades de
    comprensión lectora en los niveles <b>Literal</b>,
    <b>Inferencial</b> y <b>Crítico</b>.

    <br/><br/>

    <font color="#1565C0"><b>Promedio Final: {promedio_general} / 5.0</b></font>

    <br/><br/>

    <font color="#C79200"><b>{medalla}</b></font>

    <br/><br/>

    ✪ ¡Felicitaciones! ✪

    <br/>

    Tu dedicación, esfuerzo y amor por la lectura te permitirán
    descubrir nuevos conocimientos y alcanzar grandes sueños.
    """

    elementos.append(Paragraph(mensaje,texto))

    elementos.append(Spacer(0.1,0.1*cm))

    elementos.append(
        Paragraph(
            f"<b>Fecha de emisión:</b> {datetime.now().strftime('%d de %B de %Y')}",
            texto
        )
    )

    elementos.append(Spacer(0,0.0*cm))

    firmas = Table(
        [[
            "",
            ""
        ],
        [
            "",
            ""
        ]],
        colWidths=[0*cm,0*cm]
    )

    firmas.setStyle(TableStyle([
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("FONTSIZE",(0,0),(-1,-1),12),
        ("TOPPADDING",(0,0),(-1,-1),10)
    ]))

    elementos.append(firmas)

    elementos.append(Spacer(0,0*cm))

    pie = Table([["★★★★★   LECTURATIC   ★★★★★"]], colWidths=25*cm)

    pie.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),HexColor("#1565C0")),
        ("TEXTCOLOR",(0,0),(-1,-1),colors.white),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("FONTSIZE",(0,0),(-1,-1),14),
        ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("TOPPADDING",(0,0),(-1,-1),8)
    ]))

    elementos.append(pie)

    doc.build(elementos)

    buffer.seek(0)

    return buffer.getvalue()